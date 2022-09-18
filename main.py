import openpyxl
import os
from tkinter import *


# TODO 把列号(数字)转化为符号(str)
def get_col(col_num):
    # Excel排列顺序，A-Z，AA-AZ,BA-BZ,......
    # Excel排列顺序，1--26，27--52,53--78,......
    # 得到当前列的符号(因为处理单元格时只能使用字符串类型表示列号，不能使用整型)
    integer = int(col_num / 26)  # 取得整数
    remainder = col_num % 26  # 取得余数0-25
    temp = chr(remainder + 64)  # 取得后缀
    col_signal = ''
    if remainder == 0:
        integer -= 1
        temp = 'Z'
    if integer == 0:
        col_signal = temp
    elif integer == 1:
        col_signal = 'A' + temp
    elif integer == 2:
        col_signal = 'B' + temp
    elif integer == 3:
        col_signal = 'C' + temp
    elif integer == 4:
        col_signal = 'D' + temp
    elif integer == 5:
        col_signal = 'E' + temp
    elif integer == 6:
        col_signal = 'F' + temp
    elif integer == 7:
        col_signal = 'G' + temp
    return col_signal  # 返回当前列的符号


# TODO 搜索学分所在的行和列
def searching1():
    for s_row in range(1, max_row + 1):
        for s_col in range(1, max_column + 1):
            if ws[get_col(s_col) + str(s_row)].value == '学  分    ':
                # 返回"学分"所在的行和列
                return ws[get_col(s_col) + str(s_row)].coordinate, get_col(s_col), str(s_row), str(
                    s_col)  # 例如学分所在的行列为'C12',以上对应的值为'C12','C','12','3'


# TODO 搜索有'学分'的课的列的最大值
def searching2():
    for s2_row in range(int(search_result2), int(search_result2) + 1):
        for s2_col in range(int(search_result3) + 1, max_column + 1):
            if ws[get_col(s2_col) + str(s2_row)].value is None:
                return s2_col - 1
            else:
                if isinstance(ws[get_col(s2_col) + str(s2_row)].value, int):
                    pass
                if isinstance(ws[get_col(s2_col) + str(s2_row)].value, float):
                    pass
                else:
                    lenth = len(ws[get_col(s2_col) + str(s2_row)].value)
                    if lenth >= 3:
                        ws[get_col(s2_col) + str(s2_row)] = float(
                            ws[get_col(s2_col) + str(s2_row)].value[0] + ws[get_col(s2_col) + str(s2_row)].value[1] +
                            ws[get_col(s2_col) + str(s2_row)].value[2])
                    elif lenth == 2:  # 单元格为字符串0
                        ws[get_col(s2_col) + str(s2_row)] = float(
                            ws[get_col(s2_col) + str(s2_row)].value[0])


# TODO 将字符型的值转换为数值型
def trans():
    for t_row in range(int(search_result2) + 2, max_row + 1):
        for t_col in range(int(search_result3) + 1, search2_result + 1):  # '有'学分'的课的列的最大值'
            # TODO 如果是'- '，则跳过，否则把数字×1
            if ws[get_col(t_col) + str(t_row)].value == '- ':
                pass
            elif isinstance(ws[get_col(t_col) + str(t_row)].value, int):
                pass
            elif isinstance(ws[get_col(t_col) + str(t_row)].value, float):
                pass
            else:
                length = len(ws[get_col(t_col) + str(t_row)].value)
                if length >= 3:
                    ws[get_col(t_col) + str(t_row)] = float(
                        ws[get_col(t_col) + str(t_row)].value[0] + ws[get_col(t_col) + str(t_row)].value[1] +
                        ws[get_col(t_col) + str(t_row)].value[2] + ws[get_col(t_col) + str(t_row)].value[3])
                elif length == 2:
                    ws[get_col(t_col) + str(t_row)] = float(
                        ws[get_col(t_col) + str(t_row)].value[0] + ws[get_col(t_col) + str(t_row)].value[1])


# TODO 计算加权成绩并写入
def compute():
    for c_row in range(int(search_result2) + 2, max_row + 1):
        c_temp = 0.0000000
        mark = 0.0000000
        for c_col in range(int(search_result3) + 1, search2_result + 1):  # '最后有学号的一列'
            # TODO 如果是"-"就跳过，如果不是则进行计算
            if ws[get_col(c_col) + str(c_row)].value == '- ':
                pass
            else:
                c_temp += ws[get_col(c_col) + str(c_row)].value * ws[get_col(c_col) + str(search_result2)].value
                mark += ws[get_col(c_col) + str(search_result2)].value
        ws[get_col(max_column + 1) + str(c_row)] = c_temp / mark


# TODO 循环读入文件夹中的Excel文件
paths = "./"
for filenames in os.listdir(paths):
    if filenames.endswith('.xlsx'):
        wb = openpyxl.load_workbook(filename=filenames)  # 加载对应的文件Workbook
        ws = wb['Sheet1      ']

        max_column = ws.max_column
        max_row = ws.max_row

        # TODO 逐行逐列寻找到"学分"，并返回学分的行和列
        search_result = searching1()
        search_result0 = search_result[0]
        search_result1 = search_result[1]
        search_result2 = search_result[2]
        search_result3 = search_result[3]
        # TODO 搜索有'学分'列的最大值(因为不是所有的列都有学科)
        search2_result = searching2()

        ws[get_col(max_column + 1) + str(int(search_result2) + 1)].value = '加权成绩py'
        # TODO 把字符型(左上角有绿色角标)的成绩转换为数值型
        trans()
        # TODO 计算加权成绩并在max_column+1写入(最后一列的下一列)
        compute()

        wb.save(filename=filenames)  # 保存修改后的Excel文件
# TODO '运行完成'字样的输出模块
windows = Tk()
text = Label(windows, text="运行完成", bg="yellow", fg="red",
             font=('Times', 100))
text.pack()
windows.config(background="#bfa")
windows.geometry('600x300+500+180')
windows.mainloop()
